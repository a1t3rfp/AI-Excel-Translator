from openpyxl import load_workbook
from progress import save_progress, load_progress, clear_progress
from translator import translator
import time

TEXT_COLUMNS = [11,12,13,14,15,16,17,18]

BATCH_SIZE = 10


def translate_excel(path, callback=None):

    wb = load_workbook(path)

    ws = wb["Вопросы"]

    output = path.replace(".xlsx", "_translated.xlsx")

    total = ws.max_row - 1

    start_row = max(2, load_progress())

    row = start_row

    while row <= ws.max_row:

        batch_cells = []
        batch_texts = []

        end_row = min(row + BATCH_SIZE - 1, ws.max_row)

        # -----------------------------
        # собираем пакет
        # -----------------------------

        for r in range(row, end_row + 1):

            current_cells = []

            for col in TEXT_COLUMNS:

                cell = ws.cell(r, col)

                current_cells.append(cell)

                batch_texts.append("" if cell.value is None else str(cell.value))

            batch_cells.append(current_cells)

        # -----------------------------
        # перевод
        # -----------------------------

        translated = translator.translate_batch(batch_texts)

        index = 0

        for cells in batch_cells:

            for cell in cells:

                cell.value = translated[index]

                index += 1

        # -----------------------------
        # сохранить
        # -----------------------------

        while True:

            try:

                wb.save(output)

                break

            except PermissionError:

                print()

                print("===================================")
                print("Закрой translated.xlsx")
                print("===================================")

                time.sleep(3)

        save_progress(end_row)

        if callback:

            callback(end_row - 1, total)

        row = end_row + 1

    wb.save(output)

    clear_progress()

    return output